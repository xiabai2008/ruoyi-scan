# Excel 报告生成（openpyxl，零系统依赖）— D8.4
#
# 三 Sheet 结构：
#   Sheet 1 摘要：目标/CMS/模式/耗时/漏洞数 + 风险分布饼图
#   Sheet 2 漏洞详情：8 列表头 + 自动筛选 + 严重度整行着色 + 冻结表头
#   Sheet 3 修复建议：3 列（漏洞名称/严重度/修复建议）
import os
from datetime import datetime

from openpyxl import Workbook
from openpyxl.chart import PieChart, Reference
from openpyxl.chart.label import DataLabelList
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter

from core.models import (
    SEVERITY_CN,
    SEVERITY_HIGH,
    SEVERITY_LOW,
    SEVERITY_MEDIUM,
    STATUS_CONFIRMED,
    STATUS_SAFE,
    STATUS_UNKNOWN,
)

# 中文字体（Excel 会用系统字体渲染）
_FONT = "微软雅黑"
_FONT_MONO = "Consolas"

# 严重度整行填充色（ARGB）
_SEV_FILL = {
    SEVERITY_HIGH: PatternFill(start_color="FFF4CCCC", end_color="FFF4CCCC", fill_type="solid"),  # 浅红
    SEVERITY_MEDIUM: PatternFill(start_color="FFFFF2CC", end_color="FFFFF2CC", fill_type="solid"),  # 浅黄
    SEVERITY_LOW: PatternFill(start_color="FFD9EAD3", end_color="FFD9EAD3", fill_type="solid"),  # 浅绿
}
# 饼图分段颜色（高红/中黄/低绿）
_PIE_COLORS = ["CF222E", "D4A72C", "1A7F37"]
# 表头填充
_HEADER_FILL = PatternFill(start_color="FF0969DA", end_color="FF0969DA", fill_type="solid")
_SUBHEADER_FILL = PatternFill(start_color="FF656D76", end_color="FF656D76", fill_type="solid")
# 摘要标签列填充
_LABEL_FILL = PatternFill(start_color="FFF6F8FA", end_color="FFF6F8FA", fill_type="solid")
# 状态中文名
_STATUS_CN = {
    STATUS_CONFIRMED: "已确认",
    STATUS_SAFE: "安全",
    STATUS_UNKNOWN: "未知",
}

# 细边框
_THIN_BORDER = Border(
    left=Side(style="thin", color="D0D7DE"),
    right=Side(style="thin", color="D0D7DE"),
    top=Side(style="thin", color="D0D7DE"),
    bottom=Side(style="thin", color="D0D7DE"),
)


def _font(size=10, bold=False, color="FF1F2328", mono=False):
    """构造 Font 对象"""
    return Font(name=_FONT_MONO if mono else _FONT, size=size, bold=bold, color=color)


def _header_font(size=10, color="FFFFFFFF"):
    """表头字体（白字加粗）"""
    return Font(name=_FONT, size=size, bold=True, color=color)


def _style_header_cell(cell, fill=None, font=None):
    """设置表头单元格样式"""
    if fill:
        cell.fill = fill
    cell.font = font or _header_font()
    cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
    cell.border = _THIN_BORDER


def _style_data_cell(cell, mono=False, bold=False, color="FF1F2328", fill=None, align="left"):
    """设置数据单元格样式"""
    cell.font = _font(size=9, bold=bold, color=color, mono=mono)
    cell.alignment = Alignment(horizontal=align, vertical="top", wrap_text=True)
    cell.border = _THIN_BORDER
    if fill:
        cell.fill = fill


def _build_summary_sheet(ws, builder, gen_time):
    """Sheet 1: 摘要（目标/CMS/模式/耗时/漏洞数 + 风险分布饼图）"""
    ws.title = "摘要"
    # 列宽
    ws.column_dimensions["A"].width = 18
    ws.column_dimensions["B"].width = 40

    # 标题
    ws.merge_cells("A1:B1")
    ws["A1"] = "若依综合漏洞检测报告"
    ws["A1"].font = Font(name=_FONT, size=16, bold=True, color="FF1F2328")
    ws["A1"].alignment = Alignment(horizontal="center", vertical="center")
    ws.row_dimensions[1].height = 30

    # 摘要信息
    target = builder.target or "未指定"
    scan_time = builder.summary.get("started_at", "")
    if isinstance(scan_time, (int, float)) and scan_time:
        scan_time = datetime.fromtimestamp(scan_time).strftime("%Y-%m-%d %H:%M:%S")
    if not scan_time:
        scan_time = datetime.now().strftime("%Y-%m-%d %H:%M:%S")

    duration = builder.summary.get("duration", 0)
    req_count = builder.summary.get("request_count", 0)
    mode = builder.summary.get("mode", "")
    fp = builder.summary.get("fingerprint", {}) or {}
    fp_cms = fp.get("cms", "")
    fp_conf = fp.get("confidence", 0)

    info_rows = [
        ("扫描目标", target),
        ("CMS 识别", f"{fp_cms or '未识别'}（置信度 {fp_conf:.2f}）" if fp_cms else "未识别"),
        ("扫描模式", str(mode) if mode else "未指定"),
        ("扫描时间", str(scan_time)),
        ("报告生成", gen_time),
        ("耗时（秒）", f"{duration:.2f}" if isinstance(duration, (int, float)) else str(duration)),
        ("请求数", str(req_count)),
    ]
    for i, (k, v) in enumerate(info_rows, start=2):
        ws.cell(row=i, column=1, value=k)
        ws.cell(row=i, column=2, value=v)
        _style_data_cell(ws.cell(row=i, column=1), bold=True, fill=_LABEL_FILL)
        _style_data_cell(ws.cell(row=i, column=2))

    # 风险分布
    dist_start = len(info_rows) + 4  # 留空行
    ws.cell(row=dist_start, column=1, value="风险分布（仅统计确认漏洞）")
    ws.cell(row=dist_start, column=1).font = Font(name=_FONT, size=12, bold=True, color="FF1F2328")

    # 风险分布数据表（供饼图引用）
    dist = builder.risk_distribution()
    dist_labels = ["高危", "中危", "低危"]
    dist_values = [dist["high"], dist["medium"], dist["low"]]
    for j, (label, val) in enumerate(zip(dist_labels, dist_values), start=dist_start + 1):
        ws.cell(row=j, column=1, value=label)
        ws.cell(row=j, column=2, value=val)
        _style_data_cell(ws.cell(row=j, column=1), bold=True, fill=_LABEL_FILL)
        _style_data_cell(
            ws.cell(row=j, column=2), align="center", bold=True, color="FF" + _PIE_COLORS[j - dist_start - 1]
        )
    # 合计行
    total_row = dist_start + 1 + len(dist_labels)
    ws.cell(row=total_row, column=1, value="合计")
    ws.cell(row=total_row, column=2, value=dist["total"])
    _style_data_cell(ws.cell(row=total_row, column=1), bold=True, fill=_LABEL_FILL)
    _style_data_cell(ws.cell(row=total_row, column=2), align="center", bold=True)

    # 去重统计
    dr = builder.dedup_report()
    if dr and dr.merged_groups > 0:
        dedup_row = total_row + 2
        ws.cell(row=dedup_row, column=1, value="去重统计")
        ws.cell(row=dedup_row, column=1).font = Font(name=_FONT, size=12, bold=True, color="FF1F2328")
        ws.cell(row=dedup_row + 1, column=1, value="原始结果数")
        ws.cell(row=dedup_row + 1, column=2, value=dr.original_count)
        ws.cell(row=dedup_row + 2, column=1, value="聚合后结果数")
        ws.cell(row=dedup_row + 2, column=2, value=dr.aggregated_count)
        ws.cell(row=dedup_row + 3, column=1, value="合并组数")
        ws.cell(row=dedup_row + 3, column=2, value=dr.merged_groups)
        for r in range(dedup_row + 1, dedup_row + 4):
            _style_data_cell(ws.cell(row=r, column=1), bold=True, fill=_LABEL_FILL)
            _style_data_cell(ws.cell(row=r, column=2), align="center")

    # 饼图
    pie = PieChart()
    pie.title = "风险分布"
    labels_ref = Reference(ws, min_col=1, min_row=dist_start + 1, max_row=dist_start + len(dist_labels))
    data_ref = Reference(ws, min_col=2, min_row=dist_start, max_row=dist_start + len(dist_labels))
    pie.add_data(data_ref, titles_from_data=True)
    pie.set_categories(labels_ref)
    pie.height = 8  # cm
    pie.width = 12
    pie.dataLabels = DataLabelList(showVal=True, showCatName=True)
    # 设置分段颜色
    from openpyxl.chart.series import DataPoint

    if pie.series:
        series = pie.series[0]
        series.data_points = []
        for idx, color in enumerate(_PIE_COLORS):
            dp = DataPoint(idx=idx)
            dp.graphicalProperties.solidFill = color
            series.data_points.append(dp)
    ws.add_chart(pie, f"D{dist_start}")


def _build_vuln_sheet(ws, builder):
    """Sheet 2: 漏洞详情（8 列表头 + 自动筛选 + 严重度着色 + 冻结表头）"""
    ws.title = "漏洞详情"
    headers = ["#", "漏洞名称", "严重度", "状态", "URL", "证据", "修复建议", "命中次数"]
    for j, h in enumerate(headers, start=1):
        cell = ws.cell(row=1, column=j, value=h)
        _style_header_cell(cell, fill=_HEADER_FILL)
    ws.row_dimensions[1].height = 24

    # 列宽
    col_widths = [5, 22, 8, 10, 35, 30, 30, 10]
    for j, w in enumerate(col_widths, start=1):
        ws.column_dimensions[get_column_letter(j)].width = w

    confirmed = builder.confirmed_results()
    for i, r in enumerate(confirmed, start=2):
        sev_cn = SEVERITY_CN.get(r.severity, r.severity)
        status_cn = _STATUS_CN.get(r.status, r.status)
        hit_count = getattr(r, "hit_count", 1) if hasattr(r, "hit_count") else 1
        row_data = [i - 1, r.name, sev_cn, status_cn, r.url or "", r.evidence or "", r.fix or "", hit_count]
        row_fill = _SEV_FILL.get(r.severity)
        for j, val in enumerate(row_data, start=1):
            cell = ws.cell(row=i, column=j, value=val)
            mono = j in (5, 6)  # URL/证据列用等宽字体
            align = "center" if j in (1, 3, 4, 8) else "left"
            bold = j == 2  # 漏洞名称加粗
            _style_data_cell(cell, mono=mono, bold=bold, fill=row_fill, align=align)

    # 冻结表头
    ws.freeze_panes = "A2"
    # 自动筛选
    if confirmed:
        last_col = get_column_letter(len(headers))
        ws.auto_filter.ref = f"A1:{last_col}{len(confirmed) + 1}"

    # 其他结果（非 CONFIRMED）追加在下方，用副表头分隔
    all_results = builder._effective_results()
    others = [r for r in all_results if r.status != STATUS_CONFIRMED]
    if others:
        sep_row = len(confirmed) + 3
        ws.merge_cells(start_row=sep_row, start_column=1, end_row=sep_row, end_column=len(headers))
        sep_cell = ws.cell(row=sep_row, column=1, value="其他结果（未确认/安全）")
        sep_cell.font = Font(name=_FONT, size=11, bold=True, color="FFFFFFFF")
        sep_cell.fill = _SUBHEADER_FILL
        sep_cell.alignment = Alignment(horizontal="center", vertical="center")
        sep_cell.border = _THIN_BORDER
        ws.row_dimensions[sep_row].height = 22

        other_headers = ["#", "名称", "严重度", "状态", "URL", "证据", "", ""]
        for j, h in enumerate(other_headers, start=1):
            if h:
                cell = ws.cell(row=sep_row + 1, column=j, value=h)
                _style_header_cell(cell, fill=_SUBHEADER_FILL)

        for i, r in enumerate(others, start=sep_row + 2):
            sev_cn = SEVERITY_CN.get(r.severity, r.severity)
            status_cn = _STATUS_CN.get(r.status, r.status)
            row_data = [i - sep_row - 1, r.name, sev_cn, status_cn, r.url or "", r.evidence or "", "", ""]
            for j, val in enumerate(row_data, start=1):
                cell = ws.cell(row=i, column=j, value=val)
                mono = j in (5, 6)
                align = "center" if j in (1, 3, 4) else "left"
                _style_data_cell(cell, mono=mono, align=align)


def _build_fix_sheet(ws, builder):
    """Sheet 3: 修复建议（3 列：漏洞名称/严重度/修复建议）"""
    ws.title = "修复建议"
    headers = ["#", "漏洞名称", "严重度", "修复建议"]
    for j, h in enumerate(headers, start=1):
        cell = ws.cell(row=1, column=j, value=h)
        _style_header_cell(cell, fill=_HEADER_FILL)
    ws.row_dimensions[1].height = 24

    col_widths = [5, 25, 10, 50]
    for j, w in enumerate(col_widths, start=1):
        ws.column_dimensions[get_column_letter(j)].width = w

    confirmed = builder.confirmed_results()
    for i, r in enumerate(confirmed, start=2):
        sev_cn = SEVERITY_CN.get(r.severity, r.severity)
        row_data = [i - 1, r.name, sev_cn, r.fix or "暂无建议"]
        row_fill = _SEV_FILL.get(r.severity)
        for j, val in enumerate(row_data, start=1):
            cell = ws.cell(row=i, column=j, value=val)
            align = "center" if j in (1, 3) else "left"
            bold = j == 2
            _style_data_cell(cell, bold=bold, fill=row_fill, align=align)

    ws.freeze_panes = "A2"


def render_xlsx(builder, out_path):
    """渲染 Excel 报告到 out_path

    Args:
        builder: ReportBuilder 实例（使用 _effective_results() 获取去重后结果）
        out_path: 输出文件路径
    Returns:
        out_path
    """
    wb = Workbook()
    gen_time = datetime.now().strftime("%Y-%m-%d %H:%M:%S")

    # Sheet 1: 摘要
    ws1 = wb.active
    _build_summary_sheet(ws1, builder, gen_time)

    # Sheet 2: 漏洞详情
    ws2 = wb.create_sheet("漏洞详情")
    _build_vuln_sheet(ws2, builder)

    # Sheet 3: 修复建议
    ws3 = wb.create_sheet("修复建议")
    _build_fix_sheet(ws3, builder)

    # 确保输出目录存在
    out_dir = os.path.dirname(out_path)
    if out_dir:
        os.makedirs(out_dir, exist_ok=True)

    wb.save(out_path)
    return out_path
