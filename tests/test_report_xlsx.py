# D8.4 Excel 报告生成单元测试
import os
import sys
import tempfile

sys.path.insert(0, os.path.dirname(os.path.dirname(os.path.abspath(__file__))))

from openpyxl import load_workbook

from common.models import (ScanResult, STATUS_CONFIRMED, STATUS_SAFE, STATUS_UNKNOWN,
                          SEVERITY_HIGH, SEVERITY_MEDIUM, SEVERITY_LOW)
from core.report import ReportBuilder
from core.report_xlsx import render_xlsx


def _sample_results():
    return [
        ScanResult(kind='vuln', name='SQL注入', severity=SEVERITY_HIGH,
                   status=STATUS_CONFIRMED, url='http://x.com/sqli',
                   evidence='报错: You have an error in SQL syntax',
                   fix='使用预编译语句，禁止拼接 SQL'),
        ScanResult(kind='vuln', name='XSS', severity=SEVERITY_MEDIUM,
                   status=STATUS_CONFIRMED, url='http://x.com/xss',
                   evidence='<script>alert(1)</script>',
                   fix='对输出做 HTML 转义'),
        ScanResult(kind='info', name='端口扫描', severity=SEVERITY_LOW,
                   status=STATUS_SAFE, url='http://x.com:8080',
                   evidence='端口关闭'),
    ]


def test_xlsx_file_created():
    """Excel 文件生成成功"""
    builder = ReportBuilder(results=_sample_results(), target='http://x.com')
    with tempfile.TemporaryDirectory() as tmpdir:
        xlsx_path = os.path.join(tmpdir, 'report.xlsx')
        render_xlsx(builder, xlsx_path)
        assert os.path.exists(xlsx_path)


def test_xlsx_non_empty():
    """Excel 文件非空（>2000 字节）"""
    builder = ReportBuilder(results=_sample_results(), target='http://x.com')
    with tempfile.TemporaryDirectory() as tmpdir:
        xlsx_path = os.path.join(tmpdir, 'report.xlsx')
        render_xlsx(builder, xlsx_path)
        assert os.path.getsize(xlsx_path) > 2000


def test_xlsx_header_valid():
    """Excel 文件头部为 PK（xlsx 本质是 zip）"""
    builder = ReportBuilder(results=_sample_results(), target='http://x.com')
    with tempfile.TemporaryDirectory() as tmpdir:
        xlsx_path = os.path.join(tmpdir, 'report.xlsx')
        render_xlsx(builder, xlsx_path)
        with open(xlsx_path, 'rb') as f:
            header = f.read(2)
        assert header == b'PK'


def test_xlsx_three_sheets():
    """Excel 文件含三个 Sheet（摘要/漏洞详情/修复建议）"""
    builder = ReportBuilder(results=_sample_results(), target='http://x.com')
    with tempfile.TemporaryDirectory() as tmpdir:
        xlsx_path = os.path.join(tmpdir, 'report.xlsx')
        render_xlsx(builder, xlsx_path)
        wb = load_workbook(xlsx_path)
        assert '摘要' in wb.sheetnames, f'应含摘要 Sheet，实际 {wb.sheetnames}'
        assert '漏洞详情' in wb.sheetnames, f'应含漏洞详情 Sheet，实际 {wb.sheetnames}'
        assert '修复建议' in wb.sheetnames, f'应含修复建议 Sheet，实际 {wb.sheetnames}'
        assert len(wb.sheetnames) == 3


def test_xlsx_summary_contains_target():
    """摘要 Sheet 含目标信息"""
    builder = ReportBuilder(results=_sample_results(), target='http://x.com')
    with tempfile.TemporaryDirectory() as tmpdir:
        xlsx_path = os.path.join(tmpdir, 'report.xlsx')
        render_xlsx(builder, xlsx_path)
        wb = load_workbook(xlsx_path)
        ws = wb['摘要']
        # 在前 10 行中查找目标
        found = False
        for row in ws.iter_rows(min_row=1, max_row=10, max_col=2, values_only=True):
            if row[1] == 'http://x.com':
                found = True
                break
        assert found, '摘要 Sheet 应含目标 http://x.com'


def test_xlsx_vuln_details_content():
    """漏洞详情 Sheet 含漏洞名称和修复建议"""
    builder = ReportBuilder(results=_sample_results(), target='http://x.com')
    with tempfile.TemporaryDirectory() as tmpdir:
        xlsx_path = os.path.join(tmpdir, 'report.xlsx')
        render_xlsx(builder, xlsx_path)
        wb = load_workbook(xlsx_path)
        ws = wb['漏洞详情']
        # 收集所有单元格文本
        all_text = []
        for row in ws.iter_rows(values_only=True):
            for cell in row:
                if cell:
                    all_text.append(str(cell))
        all_content = ' '.join(all_text)
        assert 'SQL注入' in all_content, '漏洞详情应含 SQL注入'
        assert 'XSS' in all_content, '漏洞详情应含 XSS'
        assert '预编译语句' in all_content, '漏洞详情应含修复建议'


def test_xlsx_no_confirmed():
    """无确认漏洞时不报错"""
    results = [
        ScanResult(kind='info', name='安全检查', severity=SEVERITY_LOW,
                   status=STATUS_SAFE, url='http://x.com/safe', evidence='无漏洞'),
    ]
    builder = ReportBuilder(results=results, target='http://x.com')
    with tempfile.TemporaryDirectory() as tmpdir:
        xlsx_path = os.path.join(tmpdir, 'report.xlsx')
        render_xlsx(builder, xlsx_path)
        assert os.path.getsize(xlsx_path) > 0


def test_xlsx_empty_results():
    """空结果不报错"""
    builder = ReportBuilder(results=[], target='http://x.com')
    with tempfile.TemporaryDirectory() as tmpdir:
        xlsx_path = os.path.join(tmpdir, 'report.xlsx')
        render_xlsx(builder, xlsx_path)
        assert os.path.exists(xlsx_path)
        wb = load_workbook(xlsx_path)
        assert len(wb.sheetnames) == 3


def test_xlsx_with_dedup():
    """去重后 Excel 生成正确（2 条同指纹 → 1 条）"""
    extra = {'vuln_type': 'arbitrary_file_read', 'payload_class': 'traversal'}
    results = [
        ScanResult(kind='vuln', name='任意文件读取', severity=SEVERITY_HIGH,
                   status=STATUS_CONFIRMED, url='http://x.com/a?b=1',
                   evidence='root:x:0', fix='限制参数', extra=extra),
        ScanResult(kind='vuln', name='任意文件读取', severity=SEVERITY_HIGH,
                   status=STATUS_CONFIRMED, url='http://x.com/a?b=2',
                   evidence='root:x:0', fix='限制参数', extra=extra),
    ]
    builder = ReportBuilder(results=results, target='http://x.com')
    assert len(builder.confirmed_results()) == 1  # 去重后 1 条
    with tempfile.TemporaryDirectory() as tmpdir:
        xlsx_path = os.path.join(tmpdir, 'report.xlsx')
        render_xlsx(builder, xlsx_path)
        assert os.path.getsize(xlsx_path) > 2000
        wb = load_workbook(xlsx_path)
        ws = wb['摘要']
        # 验证去重统计信息
        all_text = []
        for row in ws.iter_rows(values_only=True):
            for cell in row:
                if cell:
                    all_text.append(str(cell))
        content = ' '.join(all_text)
        assert '去重统计' in content, '摘要 Sheet 应含去重统计'


def test_xlsx_freeze_panes():
    """漏洞详情 Sheet 冻结表头（A2）"""
    builder = ReportBuilder(results=_sample_results(), target='http://x.com')
    with tempfile.TemporaryDirectory() as tmpdir:
        xlsx_path = os.path.join(tmpdir, 'report.xlsx')
        render_xlsx(builder, xlsx_path)
        wb = load_workbook(xlsx_path)
        ws = wb['漏洞详情']
        assert ws.freeze_panes == 'A2', f'应冻结 A2，实际 {ws.freeze_panes}'


def test_xlsx_auto_filter():
    """漏洞详情 Sheet 设置自动筛选"""
    builder = ReportBuilder(results=_sample_results(), target='http://x.com')
    with tempfile.TemporaryDirectory() as tmpdir:
        xlsx_path = os.path.join(tmpdir, 'report.xlsx')
        render_xlsx(builder, xlsx_path)
        wb = load_workbook(xlsx_path)
        ws = wb['漏洞详情']
        assert ws.auto_filter.ref is not None, '应设置自动筛选'


def test_xlsx_with_summary():
    """带 summary 的 Excel 生成正确"""
    summary = {'started_at': 1718700000, 'duration': 12.5, 'mode': 'all',
               'request_count': 100}
    builder = ReportBuilder(results=_sample_results(), target='http://x.com', summary=summary)
    with tempfile.TemporaryDirectory() as tmpdir:
        xlsx_path = os.path.join(tmpdir, 'report.xlsx')
        render_xlsx(builder, xlsx_path)
        assert os.path.getsize(xlsx_path) > 0
        wb = load_workbook(xlsx_path)
        ws = wb['摘要']
        all_text = []
        for row in ws.iter_rows(values_only=True):
            for cell in row:
                if cell:
                    all_text.append(str(cell))
        content = ' '.join(all_text)
        assert '100' in content, '摘要应含请求数 100'


if __name__ == '__main__':
    test_xlsx_file_created()
    test_xlsx_non_empty()
    test_xlsx_header_valid()
    test_xlsx_three_sheets()
    test_xlsx_summary_contains_target()
    test_xlsx_vuln_details_content()
    test_xlsx_no_confirmed()
    test_xlsx_empty_results()
    test_xlsx_with_dedup()
    test_xlsx_freeze_panes()
    test_xlsx_auto_filter()
    test_xlsx_with_summary()
    print('All D8.4 Excel tests passed!')
